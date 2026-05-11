"""Tabular extractor — converts Excel/CSV tables into rules.

Each row becomes one rule with statement composed from header + row values.
When column headers indicate financial metadata (account codes, tax
categories), the extracted rules are tagged with that metadata.
Uses openpyxl for XLSX. See CLAUDE.md §14.11, IMPROVEMENT.md §3 提案5.
"""

from __future__ import annotations

import csv
import re
from io import StringIO
from typing import Any

from rulerepo_server.core.logging import get_logger
from rulerepo_server.services.extraction.extractors import CandidateRule, SourceFile

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# Financial metadata column detection
# ---------------------------------------------------------------------------
_ACCOUNT_CODE_HEADERS = re.compile(
    r"(?:勘定科目|account[\s_]?code|account[\s_]?number|gl[\s_]?code|科目コード|勘定コード)",
    re.IGNORECASE,
)
_TAX_CATEGORY_HEADERS = re.compile(
    r"(?:税区分|tax[\s_]?category|tax[\s_]?type|tax[\s_]?code|消費税区分|課税区分|源泉徴収)",
    re.IGNORECASE,
)
_COST_CENTER_HEADERS = re.compile(
    r"(?:コストセンター|cost[\s_]?center|部門コード|department[\s_]?code)",
    re.IGNORECASE,
)
_AMOUNT_HEADERS = re.compile(
    r"(?:金額|上限|下限|amount|limit|threshold|budget|ceiling)",
    re.IGNORECASE,
)


class TabularExtractor:
    """Extracts rules from tabular data (XLSX/CSV).

    Each row produces a rule statement combining column headers
    with cell values into a natural-language statement.

    When financial metadata columns are detected (account codes,
    tax categories, cost centers), the extracted rules are enriched
    with ``financial_metadata`` in ``source_refs`` and tagged
    accordingly.
    """

    source_types = ["spreadsheet", "csv"]

    async def extract(self, source: SourceFile) -> list[CandidateRule]:
        """Extract candidate rules from a tabular file.

        Args:
            source: The tabular source file.

        Returns:
            One CandidateRule per data row.
        """
        logger.info("tabular_extraction_started", path=str(source.path))

        suffix = source.path.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            rows = self._read_xlsx(source)
        elif suffix == ".csv" or source.content:
            rows = self._read_csv(source)
        else:
            logger.warning("tabular_unsupported_format", suffix=suffix)
            return []

        # Detect financial columns from headers
        all_headers: set[str] = set()
        for row in rows:
            all_headers.update(k for k in row if not k.startswith("_"))
        fin_columns = _detect_financial_columns(all_headers)
        has_financial = bool(fin_columns)

        candidates: list[CandidateRule] = []
        for row in rows:
            statement = self._row_to_statement(row, source.metadata)
            if not statement:
                continue

            source_refs: dict[str, Any] = {
                "document": str(source.path),
                "sheet": source.metadata.get("sheet", ""),
                "row": row.get("_row_number", 0),
            }

            # Attach financial metadata when detected
            if has_financial:
                fin_meta = _extract_financial_metadata(row, fin_columns)
                if fin_meta:
                    source_refs["financial_metadata"] = fin_meta

            tags = ["tabular", "extracted"]
            if has_financial:
                tags.append("financial")

            candidates.append(
                CandidateRule(
                    statement=statement,
                    modality=row.get("modality", "MUST"),
                    severity=row.get("severity", "MEDIUM"),
                    scope=source.metadata.get("scope", []),
                    source_refs=source_refs,
                    department=source.metadata.get("department", "finance"),
                    tags=tags,
                    applicable_subject_kinds=["transaction", "event"],
                    confidence=0.8,
                )
            )

        logger.info(
            "tabular_extraction_complete",
            candidates=len(candidates),
            financial_columns=list(fin_columns.keys()) if fin_columns else [],
        )
        return candidates

    def _read_xlsx(self, source: SourceFile) -> list[dict[str, Any]]:
        """Read rows from an XLSX file using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl_not_installed")
            return []

        wb = openpyxl.load_workbook(source.path, read_only=True)
        sheet_name = source.metadata.get("sheet")
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        rows: list[dict[str, Any]] = []
        headers: list[str] = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(cell or f"col_{j}") for j, cell in enumerate(row)]
                continue
            row_dict = {headers[j]: cell for j, cell in enumerate(row) if j < len(headers)}
            row_dict["_row_number"] = i + 1
            rows.append(row_dict)

        wb.close()
        return rows

    def _read_csv(self, source: SourceFile) -> list[dict[str, Any]]:
        """Read rows from a CSV file or content string."""
        content = source.content or ""
        if not content and source.path.exists():
            content = source.path.read_text(encoding="utf-8", errors="replace")

        reader = csv.DictReader(StringIO(content))
        rows: list[dict[str, Any]] = []
        for i, row in enumerate(reader):
            row["_row_number"] = i + 2  # 1-indexed, skip header
            rows.append(dict(row))
        return rows

    def _row_to_statement(self, row: dict[str, Any], metadata: dict[str, Any]) -> str:
        """Convert a data row to a natural-language rule statement."""
        # Skip empty rows
        values = {k: v for k, v in row.items() if v and not k.startswith("_")}
        if not values:
            return ""

        # If there's a 'statement' column, use it directly
        if "statement" in values:
            return str(values["statement"])

        # Otherwise compose from header-value pairs
        parts = [f"{k}: {v}" for k, v in values.items() if k not in ("modality", "severity")]
        template = metadata.get("statement_template", "{}")
        if template != "{}":
            return template.format(**values)

        return "; ".join(parts)


# ---------------------------------------------------------------------------
# Financial column detection and metadata extraction
# ---------------------------------------------------------------------------


def _detect_financial_columns(headers: set[str]) -> dict[str, str]:
    """Detect which headers represent financial metadata.

    Returns:
        A mapping from header name to financial metadata type
        (e.g., ``{"勘定科目": "account_code", "税区分": "tax_category"}``).
    """
    result: dict[str, str] = {}
    for header in headers:
        if _ACCOUNT_CODE_HEADERS.search(header):
            result[header] = "account_code"
        elif _TAX_CATEGORY_HEADERS.search(header):
            result[header] = "tax_category"
        elif _COST_CENTER_HEADERS.search(header):
            result[header] = "cost_center"
        elif _AMOUNT_HEADERS.search(header):
            result[header] = "amount_limit"
    return result


def _extract_financial_metadata(
    row: dict[str, Any],
    fin_columns: dict[str, str],
) -> dict[str, str]:
    """Extract financial metadata values from a data row.

    Args:
        row: A single data row.
        fin_columns: Mapping from header to financial type.

    Returns:
        A dict like ``{"account_code": "4110", "tax_category": "課税"}``.
    """
    meta: dict[str, str] = {}
    for header, fin_type in fin_columns.items():
        value = row.get(header)
        if value is not None and str(value).strip():
            meta[fin_type] = str(value).strip()
    return meta
